import openpyxl
from django.http import HttpResponse
from .models import Author, Book, BorrowRecord

def export_to_excel():
    wb = openpyxl.Workbook()
    
    # Authors sheet
    ws1 = wb.active
    ws1.title = "Authors"
    ws1.append(["ID", "Name", "Email", "Bio"])
    for author in Author.objects.all():
        ws1.append([author.id, author.name, author.email, author.bio])

    # Books sheet
    ws2 = wb.create_sheet(title="Books")
    ws2.append(["ID", "Title", "Genre", "Published Date", "Author"])
    for book in Book.objects.all():
        ws2.append([book.id, book.title, book.genre, book.published_date, book.author.name])

    # BorrowRecords sheet
    ws3 = wb.create_sheet(title="BorrowRecords")
    ws3.append(["ID", "User Name", "Book", "Borrow Date", "Return Date"])
    for record in BorrowRecord.objects.all():
        ws3.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
    wb.save(response)
    return response
